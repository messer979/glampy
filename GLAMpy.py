##ensure oracle_client.bat is run first or the PATH variable is set up

import cx_Oracle, csv, time
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from datetime import datetime

def open_db_connection(auto_sheet):
    connection = None
    print("Connecting to DB. This may take a moment...")
    while connection == None:
        try:
            connection = cx_Oracle.connect(auto_sheet['B2'].value, auto_sheet['B3'].value, f"{auto_sheet['B4'].value}:{auto_sheet['B5'].value}/{auto_sheet['B6'].value}")
        except(NameError, cx_Oracle.OperationalError):
            print("Retrying connection")
            continue
        except cx_Oracle.DatabaseError:
            print(r'''Error. Either you are not on the same network as the DB or Oracle Instant Client not found. Please follow instructions sheet to ensure
                the instant client library is installed and set in Path. If you are not following the PATH
                method, ensure the up to date visual basic libraries are installed and open a command prompt session.
                Run SET PATH=C:\(path to instant client)\instantclient_18_5;%PATH% then run the executable.
                ''')
            input('Please close window.')
    print(f"Connected to {auto_sheet['B4'].value}:{auto_sheet['B5'].value}/{auto_sheet['B6'].value}")
    return connection

def find_date_row(col, working_date):
        for cell in temp_raw.iter_rows(min_row=11, min_col=col, max_col=col):
            if cell[0].value == working_date:
                return cell[0].row

def update_raw_table(raw_table, col_update, update_row):
    for i in range(0,len(col_update)):
        try:
            if col_update[i] == '?':
                continue
            if update_row == None:
                continue
            raw_table[col_update[i] + str(update_row)].value = qrow[i+1]
        except IndexError:
            continue

    ## save as new spreadsheet
def save_output():
    while True:
        try:
            output_file = f"FSC_GLAM_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            wb.save(output_file)
            wb.save('GLAM Master Report.xlsx')
        except PermissionError:
            print('Unable to save spreadhseet. Please ensure sheet is closed.')
            resume = input('Retry? (y/n)\n')
            if resume.upper() == 'Y':
                continue
            elif resume.upper() == 'N':
                print('Sheet not saved')
                break
        break


def run_health_checks():
    try:
        wb = load_workbook('Data Integrity Checks.xlsx')
        wm_db_conn = open_db_connection(auto_sheet)
        ic_cursor = wm_db_conn.cursor()
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(min_col=4, max_col=7, min_row=6):
                row[2].value = ''
                row[3].value = ''
                if row[0].value == None:
                    continue
                ic_cursor = ic_cursor.execute(row[0].value.lstrip("'").replace(";"," "))
                ic_result = ic_cursor.fetchall()
                if len(ic_result) == 0:
                    row[2].value = 'P'
                elif ic_result[0][0] == 0 and len(ic_result) == 1:
                    row[2].value = 'P'
                else:
                    row[2].value = 'F'
                    row[3].value = ''
                    cell_output = [str([head[0] for head in ic_cursor.description])]
                    cell_output = cell_output + [row for row in ic_result]
                    for line in cell_output:
                        row[3].value = row[3].value + str(line) + '\n'
        save_health_checks()
    except:
        print(row[0].value.lstrip("'").replace(";"," "))
        raise

def save_health_checks():
    while True:
        try:
            wb.save('Data Integrity Checks.xlsx')
        except PermissionError:
            print('Unable to save spreadhseet. Please ensure sheet is closed.')
            resume = input('Retry? (y/n)\n')
            if resume.upper() == 'Y':
                continue
            elif resume.upper() == 'N':
                print('Sheet not saved')
                break
        break


print('Starting AutoGlam')

try:
    #main wb load
    wb = load_workbook('GLAM Master Report.xlsx')
    auto_sheet = wb['AutoPopulateParameters']
    raw_table = wb['Raw Table']

    ##used for date matching
    temp_book = load_workbook('GLAM Master Report.xlsx', data_only=True)
    temp_raw = temp_book['Raw Table']
except:
    input('Workbooks not found or failed to load. Please ensure GLAM Master Report.xlsx exists in the same directory as this program.')
    exit()
try:
    wm_db_conn = open_db_connection(auto_sheet)
except:
    input('Connection to DB failed. Please ensure you are on same network and any required host files are set up.')
    exit()

print('Running GLAM Queries.')
glam_cursor = wm_db_conn.cursor()
for row in auto_sheet.iter_rows(min_row=9, min_col=4):
    col_identifier = column_index_from_string(row[2].value)
    col_update = row[3].value.split(',')
    glam_cursor = glam_cursor.execute(row[1].value)
    glam_result = glam_cursor.fetchall()
    for qrow in glam_result:
        try:
            working_date = datetime.strptime(qrow[0],'%Y-%m-%d')
        except TypeError:
            print('Error encountered: ')
            print(row[1])
            print(row[1].value)
            print(qrow)
        update_row = find_date_row(col_identifier, working_date)
        update_raw_table(raw_table, col_update, update_row)

save_output()
wm_db_conn.close()
wb.close()
print("GLAM Queries run successfully.")

print("Starting health checks...")
run_health_checks()
print("Health Checks run successfully.")

print('\nAutoGlam run successfully.')
time.sleep(10)